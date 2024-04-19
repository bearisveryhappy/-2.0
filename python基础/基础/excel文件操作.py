# -*-codeing=utf-8-*-
# @Time : 2022/4/30 17:54
# @Author:熊金海
# @File : excel文件操作.py
# @Software: PyCharm

import openpyxl
import xlwt

# workbook = openpyxl.load_workbook("豆瓣排名250.xlsx")
# print(workbook.sheetnames)

workbook = xlwt.Workbook(encoding="utf-8")
sheet = workbook.add_sheet("第一张表",cell_overwrite_ok=True)
sheet.write(0,0,"你好啊！！")
workbook.save("./呵呵哒.xls")

